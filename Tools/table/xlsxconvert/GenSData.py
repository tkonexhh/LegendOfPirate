#!/bin/python
#
# coding=utf-8
import os
import sys

import codecs
from Cheetah.Template import Template
from openpyxl import load_workbook
from openpyxl.compat import range
from openpyxl.cell import get_column_letter
import inspect
import os
import Config


def script_path():
    this_file = inspect.getfile(inspect.currentframe())
    return os.path.abspath(os.path.dirname(this_file))


def format_class_name(src_class_name):
    class_name = src_class_name.capitalize()
    ret = ""
    i = 0
    while(i < len(class_name)):
        if (class_name[i] == '_'):
            if (i + 1 < len(class_name)):
                ret = ret + class_name[i + 1].upper()
            i = i + 2
        else:
            ret = ret + class_name[i]
            i = i + 1
    return ret


class XlsxParser:

    def __init__(self, belong):
        self.ClassName = None
        self.ClassMembers = []
        self.MembersComments = []
        self.filedBelongs = {}
        self.xlsxSheetData = None
        self.belong = belong
        self._filePath = ""
        self.keyType = "int"
        self.keyName = "id"
        self.clientFieldCount = 0
        self.serverFieldCount = 0

    def ParseCustomTypeForCSharp(self, typestr):
        if (typestr.find("Lint") != -1):
            size = int(typestr[4:])
            return {"Type": "List<int>", "Size": size}
        elif (typestr.find("Lbool") != -1):
            size = int(typestr[5:])
            return {"Type": "List<bool>", "Size": size}
        elif (typestr.find("Lfloat") != -1):
            size = int(typestr[6:])
            return {"Type": "List<float>", "Size": size}
        elif (typestr.find("Lstring") != -1):
            size = int(typestr[7:])
            return {"Type": "List<string>", "Size": size}
        else:
            return None

    def ParseCustomType(self, typestr):
        if (typestr.find("Lint") != -1):
            size = int(typestr[4:])
            return {"Type": "int", "Size": size}
        elif (typestr.find("Lbool") != -1):
            size = int(typestr[5:])
            return {"Type": "bool", "Size": size}
        elif (typestr.find("Lfloat") != -1):
            size = int(typestr[6:])
            return {"Type": "float", "Size": size}
        elif (typestr.find("Lstring") != -1):
            size = int(typestr[7:])
            return {"Type": "string", "Size": size}
        else:
            return None

    def CheckIsBaseType(self, cType):
        cellType = cType.strip().lower()
        if (cellType == "int" or cellType == "bool" or cellType == "string"
                or cellType == "float" or cellType == "byte"
                or cellType == "vector2" or cellType == "vector3"):
            return True
        return False

    def _checkInvalid(self, cellData):
        if (cellData.value == None or type(cellData.value) == None or (type(cellData.value) == "str" and cellData.value.strip() == "")):
            return False
        else:
            return True

    def _checkTypeInvalid(self, cellType):
        if (self.CheckIsBaseType(cellType)):
            return True
        else:
            if (self.ParseCustomType(cellType.strip()) != None):
                return True
            else:
                return False

    def GetCustomType(self):
        ret = []
        for member in self.ClassMembers:
            Type = member["Type"]
            Name = member["Name"]
            if (not self.CheckIsBaseType(Type)):
                otherType = self.ParseCustomType(Type)
                ret.append({"CustomType": Name, "ElementType": otherType[
                           "Type"],  "Size": otherType["Size"]})
        return ret

    def _checkBelongInvalid(self, cellData):
        cellVal = cellData.value.strip()
        if (cellVal == "A" or cellVal == "C" or cellVal == "N" or cellVal == "S" or cellVal == "K"):
            return True
        else:
            return False

    def _parseFiledBelong(self):
        #print ("sheet Max Column:%d" %(self.xlsxSheetData.max_column))
        for c in range(1, self.xlsxSheetData.max_column + 1):
            cellData = self.xlsxSheetData.cell(row=Config.belongLine, column=c)
            if (cellData.value != None):
                #print (len(cellData.value), cellData.value)
                if (len(cellData.value.strip()) > 1):
                    print(("ERROR: Data:Invalid col = %s row = %d must be \"K\" or \"A\" or \"C\" or \"S\" or \"N\" in FilePath: %s" % (
                        get_column_letter(c), Config.belongLine, self._filePath)))
                    assert(False)
                if (self._checkBelongInvalid(cellData)):
                    self.filedBelongs[c] = cellData.value.strip()
                    cellVal = cellData.value.strip()
                    if (cellVal == "A" or cellVal == "C" or cellVal == "K"):
                        self.clientFieldCount = self.clientFieldCount + 1
                    elif (cellVal == "A" or cellVal == "S" or cellVal == "K"):
                        self.serverFieldCount = self.serverFieldCount + 1
                else:
                    print(("ERROR: Data:column = %s row = %d Data Invalid in FilePath: %s" % (
                        get_column_letter(c), Config.belongLine, self._filePath)))
                    assert(False)
            else:
                self.filedBelongs[c] = 'N'
        return

    def _checkNeedWrite(self, col):
        belong = self.belong
        if (self.filedBelongs.has_key(col)):
            belongVal = self.filedBelongs[col]
            if (belong == "C"):
                return belongVal == "A" or belongVal == "C" or belongVal == "K"
            else:
                return belongVal == "A" or belongVal == "S" or belongVal == "K"
        else:
            print(("ERROR Not Find %d Invalid in FilePath: %s" %
                   (get_column_letter(col), self._filePath)))
            assert(False)
        return False

    def _parseMembersType(self):
        membersType = []
        membersColumKey = []
        for c in range(1, self.xlsxSheetData.max_column + 1):
            cellData = self.xlsxSheetData.cell(row=Config.typeLine, column=c)
            if (self._checkInvalid(cellData)):
                if (self._checkTypeInvalid(cellData.value)):
                    if (self._checkNeedWrite(c)):
                        membersType.append(cellData.value.strip())
                        membersColumKey.append(c)
                else:
                    print(("ERROR Type:column = %s row = 2 DataType Invalid in FilePath: %s" % (
                        get_column_letter(c), self._filePath)))
                    assert(False)
        return membersType, membersColumKey

    def _parseKeyType(self):
        cellData = self.xlsxSheetData.cell(row=Config.typeLine, column=1)
        if (cellData.value != None):
            self.keyType = cellData.value.lower()
        cellData = self.xlsxSheetData.cell(row=Config.filedNameLine, column = 1)
        self.keyName = Config.FormatGetPropName(cellData.value)

        for c in range(1, self.xlsxSheetData.max_column + 1):
            if (self.filedBelongs[c] == "K"):
                cellData = self.xlsxSheetData.cell(row=Config.filedNameLine, column = c)
                self.keyName = Config.FormatGetPropName(cellData.value)
                cellData = self.xlsxSheetData.cell(row=Config.typeLine, column = c)
                if (cellData.value != None):
                    self.keyType = cellData.value.lower()
                break

    def _getMembersName(self, col):
        cellData = self.xlsxSheetData.cell(
            row=Config.filedNameLine, column=col)
        if (self._checkInvalid(cellData)):
            return cellData.value
        else:
            print(("ERROR:column = %s row = 3 Data Invalid in FilePath: %s" %
                   (get_column_letter(col), self._filePath)))
            assert(False)

    def _parseMembersComments(self):
        for c in range(1, self.xlsxSheetData.max_column + 1):
            if (self._checkNeedWrite(c)):
                cellData = self.xlsxSheetData.cell(
                    row=Config.commentLine, column=c)
                if (cellData.value == None):
                    self.MembersComments.append("")
                else:
                    self.MembersComments.append(("%s" % (cellData.value)))
        return

    def _parseClassName(self, filePath):
        baseName = os.path.basename(filePath)
        fileName, ext = os.path.splitext(baseName)
        return fileName

    def _parseClassMembers(self):
        membersType, membersColumKey = self._parseMembersType()

        column = 0
        for memberType in membersType:
            member = {}
            if memberType.strip().lower() == "byte":
                member["Type"] = "int"
            elif memberType.strip().lower() == "vector2":
                member["Type"] = "Vector2"
            elif memberType.strip().lower() == "vector3":
                member["Type"] = "Vector3"
            else:
                member["Type"] = memberType.strip().lower()

            member["Name"] = self._getMembersName(membersColumKey[column])
            column = column + 1
            self.ClassMembers.append(member)
        return

    def _getSheetByIndex(self, workbook, index):
        sheetNames = workbook.get_sheet_names()
        #print (sheetNames)
        sheetName = sheetNames[index]
        return workbook.get_sheet_by_name(sheetName)

    def has_field(self, field_key):
        for i in range(len(self.ClassMembers)):
            if (self.ClassMembers[i]["Name"] == field_key):
                return True
        return False

    def ParseXlsx(self, filePath, is_format=False):
        self._filePath = filePath
        self.ClassName = self._parseClassName(filePath)
        self.fileBaseName = self.ClassName
        if (is_format):
            self.ClassName = format_class_name(self.ClassName)
        wb = load_workbook(filename=os.path.abspath(filePath))
        # default sheet1
        self.xlsxSheetData = self._getSheetByIndex(wb, 0)
        self._parseFiledBelong()
        self._parseClassMembers()
        self._parseMembersComments()
        self._parseKeyType()


def TemplateToFile(templateFile, xlsxParser, outFilePath, isBom=False):
    tmplFd = open(os.path.abspath(templateFile), 'r')
    strTemplate = Template(source=tmplFd.read(), searchList=[{"XlsxParser": xlsxParser},
                                                             {"TabelKeyType": xlsxParser.keyType},
                                                             {"TabelKeyName": xlsxParser.keyName},
                                                             {"UpperClassName": xlsxParser.ClassName.upper(
                                                             )},
                                                             {"FileBaseName": xlsxParser.fileBaseName},
                                                             {"ClassName": xlsxParser.ClassName},
                                                             {"ClassMembers": xlsxParser.ClassMembers},
                                                             {"MembersComments": xlsxParser.MembersComments},
                                                             {"Config": Config}])
    fd = codecs.open(outFilePath, 'w', "utf-8")
    if (isBom):
        fd.write(codecs.BOM_UTF8)
    fd.write(str(strTemplate))
    fd.close
    tmplFd.close()


class TemplateTableParser():

    def __init__(self, outDir):
        self.outFileDir = outDir

    def TemplateToFile(self):
        tmplFd = open(os.path.abspath("./csharp/TableParser.tmpl"), 'r')
        strTemplate = Template(source=tmplFd.read(),
                               searchList=[{"Config": Config}])
        fd = codecs.open(os.path.abspath("%s/TableParser.cs" %
                                         (self.outFileDir)), 'w', "utf-8")
        if (True):
            fd.write(codecs.BOM_UTF8)
        fd.write(str(strTemplate))
        fd.close
        tmplFd.close()


class CplusplusGenerator:

    def __init__(self):
        self.xlsxParser = XlsxParser("S")
        self._filePath = ""

    def Generate(self, filePath):
        self._filePath = filePath
        self.xlsxParser.ParseXlsx(filePath)

    def GenerateToFile(self, filePath, outFileDir):
        self._filePath = filePath
        self.xlsxParser.ParseXlsx(filePath)
        if (self.xlsxParser.serverFieldCount <= 0):
            print ("Ingore %s, Server Field Count Empty" % (filePath))
            return

        TemplateToFile("./cplusplus/DataItem_h.tmpl", self.xlsxParser,
                       os.path.abspath("%s/%sItem.h" % (outFileDir, self.xlsxParser.ClassName)))
        TemplateToFile("./cplusplus/DataItem_cpp.tmpl", self.xlsxParser,
                       os.path.abspath("%s/%sItem.cpp" % (outFileDir, self.xlsxParser.ClassName)))
        TemplateToFile("./cplusplus/DataTable_h.tmpl", self.xlsxParser,
                       os.path.abspath("%s/%sTable.h" % (outFileDir, self.xlsxParser.ClassName)))
        TemplateToFile("./cplusplus/DataTable_cpp.tmpl", self.xlsxParser,
                       os.path.abspath("%s/%sTable.cpp" % (outFileDir, self.xlsxParser.ClassName)))

    def UpperFirstChar(self, name):
        return name[0].upper() + name[1:]


class CsharpGenerator:

    def __init__(self):
        self.xlsxParser = XlsxParser("C")

    def Generate(self, filePath):
        self.xlsxParser.ParseXlsx(filePath, True)

    def GenerateToFile(self, filePath, gendir, extenddir):

        self.xlsxParser.ParseXlsx(filePath, True)
        if (self.xlsxParser.clientFieldCount <= 0):
            print ("Ingore %s, Client Field Count Empty" % (filePath))
            return

        if (not self.xlsxParser.has_field("Id")):
            print ("Ingore %s, Has't Id Filed" % (filePath))
            return
        templateDataFile = "./csharp/Data.tmpl"
        templateDataTableFile = "./csharp/DataTable.tmpl"
        TemplateToFile(templateDataFile, self.xlsxParser, os.path.abspath(
            "%s/TD%s.cs" % (gendir, self.xlsxParser.ClassName)), True)
        TemplateToFile(templateDataTableFile, self.xlsxParser, os.path.abspath(
            "%s/TD%sTable.cs" % (gendir, self.xlsxParser.ClassName)), True)
        templateDataFile = "./csharp/DataExtend.tmpl"
        templateDataTableFile = "./csharp/DataTableExtend.tmpl"
        TemplateToFile(templateDataFile, self.xlsxParser, os.path.abspath(
            "%s/TD%sExtend.cs" % (extenddir, self.xlsxParser.ClassName)), True)
        TemplateToFile(templateDataTableFile, self.xlsxParser, os.path.abspath(
            "%s/TD%sTableExtend.cs" % (extenddir, self.xlsxParser.ClassName)), True)

        #TemplateToFile("./csharp/TableParser.tmpl", self.xlsxParser, os.path.abspath("%s/TableParser.cs" %(outFileDir)), True)

    def UpperFirstChar(self, name):
        return name[0].upper() + name[1:]
